from openpyxl import load_workbook
wb = load_workbook("E1S2.xlsx")
f= open("models.py","r").read()
g= open("models.py","w")
g.write(f)
print(f)
for sheet in wb.sheetnames:
    items = [[col.value for col in row] for row in wb[sheet].rows]
    print(items[0])
    g.write(f"class {sheet}(models.Model):\n")
    g.write(f"""    {("".join(s+"_" for s in items[0][0].split())[:-1])}= models.CharField(max_length=10)
    {("".join(s+"_" for s in items[0][1].split()))[:-1]}= models.CharField(max_length=50)
    {("".join(s+"_" for s in items[0][2].split()))[:-1]}= models.CharField(max_length=2)
    {("".join(s+"_" for s in items[0][3].split()))[:-1]} = models.CharField(max_length=2)
    {("".join(s+"_" for s in items[0][4].split()))[:-1]}= models.CharField(max_length=2)
    {("".join(s+"_" for s in items[0][5].split()))[:-1]} = models.CharField(max_length=2)
    {("".join(s+"_" for s in items[0][6].split()))[:-1]} = models.CharField(max_length=2)
    {("".join(s+"_" for s in items[0][7].split()))[:-1]}= models.CharField(max_length=2)
    {("".join(s+"_" for s in items[0][8].split()))[:-1]} = models.CharField(max_length=2)
    {("".join(s+"_" for s in items[0][9].split()))[:-1]} = models.CharField(max_length=2)\n\n\n"""
        )